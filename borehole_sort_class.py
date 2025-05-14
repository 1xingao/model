import pandas as pd
from collections import Counter
from copy import deepcopy
from borehole_data import boreholeData
from openpyxl import Workbook
from openpyxl.styles import PatternFill

class boreholeProcessor:
    def __init__(self, boreholes, boreholes_thickness):
        self.boreholes = boreholes #地层
        self.boreholes_thickness = boreholes_thickness #地层厚度
        self.unified_sequence = [] #统一地层序列
        self.standardized_boreholes = {} #标准化后的钻孔层序
        self.preconditioning_completes_borehole = {} #预处理完成的钻孔数据


    def sort_by_standard_order(self, data_list, standard_order):
        order_map = {value: index for index, value in enumerate(standard_order)}
        sorted_list = sorted(data_list, key=lambda x: order_map.get(x, len(standard_order)))
        return sorted_list

    def get_unified_sequence(self):
        all_layers = set(layer for layers in self.boreholes.values() for layer in layers)
        all_layers = sorted(list(all_layers))
        layer_index = {layer: i for i, layer in enumerate(all_layers)}

        C = [[0 for _ in all_layers] for _ in all_layers]

        for bh_layers in self.boreholes.values():
            for i in range(len(bh_layers)):
                for j in range(len(bh_layers)):
                    if i == j:
                        continue
                    li, lj = bh_layers[i], bh_layers[j]
                    idx_i, idx_j = layer_index[li], layer_index[lj]
                    if i < j:
                        C[idx_i][idx_j] += 1
                    else:
                        C[idx_i][idx_j] -= 1

        priority = {layer: 0 for layer in all_layers}
        for i, li in enumerate(all_layers):
            for j, lj in enumerate(all_layers):
                priority[li] += C[i][j]

        sorted_layers = sorted(priority.items(), key=lambda x: -x[1])
        unified_sequence = [layer for layer, _ in sorted_layers]

        print("统一地层层序：")
        print(unified_sequence)
        self.unified_sequence = unified_sequence
        return unified_sequence

    def make_max_compatible_boreholes(self):
        error_list = []
        borehole_keys = list(self.boreholes.keys())
        boreholes_matrix = list(self.standardized_boreholes.values())
        
        for i in range(len(self.unified_sequence)-1):
            error_layer = set()
            
            for idx, borehole in enumerate(boreholes_matrix):
                start = i
                while start < len(borehole) and self.standard_names(borehole[start]) != self.unified_sequence[i]:
                    start += 1
                end = start + 1
                while end < len(borehole) and self.standard_names(borehole[end]) != self.unified_sequence[i+1]:
                    end += 1
                
                if end - start != 1:
                    for j in range(start+1, end):
                        if j >= len(borehole): break
                        error_layer.add(borehole[j])
                        
            error_list.append(self.sort_by_standard_order(list(error_layer), self.unified_sequence) if len(error_layer) > 0 else [])
        
        end_error_list = []
       
        for borehole in boreholes_matrix:
            index = len(borehole)-1
            while self.standard_names(borehole[index]) != self.unified_sequence[-1]:
                index -= 1
            if index + 1 < len(borehole):
                end_error_list.extend(borehole[index+1:])
        
        error_list.append(self.sort_by_standard_order(list(set(end_error_list)), self.unified_sequence))
        
        for i in range(len(error_list)):
            for j in range(len(error_list[i])):
                error_list[i][j] += "-"

        print(f"所有钻孔标准地层之间存在的异常层：")
        print(error_list)

        res = {}
        for idx, borehole in enumerate(boreholes_matrix):
            unified_sequence_index = 0
            temp = []
            curr_layer_index = 0
            while curr_layer_index < len(borehole):
                if unified_sequence_index >= len(self.unified_sequence):
                    temp.append(borehole[curr_layer_index])
                    curr_layer_index += 1
                    continue
                if self.standard_names(borehole[curr_layer_index]) == self.unified_sequence[unified_sequence_index]:
                    temp.append(borehole[curr_layer_index])
                    curr_layer_index += 1
                    for error_layer in error_list[unified_sequence_index]:
                        if curr_layer_index < len(borehole) and self.standard_names(borehole[curr_layer_index]) in error_layer:
                            temp.append(borehole[curr_layer_index])
                            curr_layer_index += 1
                        else:
                            temp.append(error_layer)
                unified_sequence_index += 1

            res[borehole_keys[idx]] = temp
        return res

    def print_pandas(self, standardized_boreholes):
        max_len = max(len(layers) for layers in standardized_boreholes.values())
        for bh in standardized_boreholes:
            length_diff = max_len - len(standardized_boreholes[bh])
            standardized_boreholes[bh] += ['一一'] * length_diff  
        df = pd.DataFrame(standardized_boreholes)
        print(df)

    def standard_names(self, layer_name):
        if "-" in layer_name or "0" in layer_name:
            return layer_name[:-1]
        return layer_name

    def complete_missing_layers(self):
        standardized_boreholes = {}
        problematic_boreholes = []

        for bh_name, bh_layers in self.boreholes.items():
            fixed_layers = []
            seen_layers = set()
            current_index = 0

            for layer in bh_layers:
                while current_index < len(self.unified_sequence) and self.unified_sequence[current_index] != layer and current_index < self.unified_sequence.index(layer):
                    missing = self.unified_sequence[current_index]
                    fixed_layers.append(missing + "0")
                    seen_layers.add(missing + "0")
                    current_index += 1

                if current_index < len(self.unified_sequence) and self.unified_sequence[current_index] == layer:
                    fixed_layers.append(layer)
                    seen_layers.add(layer)
                    current_index += 1
                else:
                    fixed_layers.append(layer)

            for i in range(current_index, len(self.unified_sequence)):
                fixed_layers.append(f"{self.unified_sequence[i]}0")

            if len(fixed_layers) != len(self.unified_sequence):
                problematic_boreholes.append((bh_name, "地层重复"))

            standardized_boreholes[bh_name] = fixed_layers

        print("\n标准化后的钻孔层序（倒置转重复后）：")
        self.print_pandas(deepcopy(standardized_boreholes))

        if problematic_boreholes:
            print("\n以下钻孔存在问题（已转化为地层重复）：")
            for bh, err in problematic_boreholes:
                print(f"{bh} - {err}")
        else:
            print("\n所有钻孔层序正常")

        return standardized_boreholes

    def zero_thickness_filling(self):
        res = {}
        for key, value in self.preconditioning_completes_borehole.items():
            thickness_layer = []
            layer_index = 0
            for i in range(len(value)):
                if "0" in value[i] or "-" in value[i]:
                    thickness_layer.append(0)
                else:
                    thickness_layer.append(self.boreholes_thickness[key][layer_index])
                    layer_index += 1
            res[key] = thickness_layer
        return res

    def create_boreholeing_class(self, location):
        borehole_class_list = []
        for key, value in self.boreholes.items():
            temp_borehole = boreholeData(key, location[key], value, self.boreholes_thickness[key])
            borehole_class_list.append(temp_borehole)
        return borehole_class_list

    def write_to_excel(self, file_path):
        # 创建一个新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Borehole Data"

        # 获取地层名称和钻孔名称
        layers = list(self.preconditioning_completes_borehole.values())
        borehole_names = list(self.preconditioning_completes_borehole.keys())
        thickness_data = list(self.boreholes_thickness.values())

        # 写入表头
        ws.cell(row=1, column=1, value="地层名称")
        for col, borehole_name in enumerate(borehole_names, start=2):
            ws.cell(row=1, column=col, value=f"{borehole_name} (厚度)")

        # 定义颜色填充
        fill_zero = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色
        fill_dash = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色

        # 写入地层名称和数据
        max_layers = max(len(layer) for layer in layers)
        for row in range(max_layers):
            layer_replace = layers[0][row].replace("-", "")
            layer_replace = layer_replace.replace("0", "")
            if row < len(layers[0]):
                ws.cell(row=row + 2, column=1, value=layer_replace)  # 地层名称
            for col, borehole_name in enumerate(borehole_names, start=2):
                if row < len(layers[col - 2]):
                    cell = ws.cell(row=row + 2, column=col)
                    # 根据内容设置颜色
                    if "0" in layers[col - 2][row]:
                        cell.fill = fill_zero
                    elif "-" in layers[col - 2][row]:
                        cell.fill = fill_dash
                    else:
                        cell.fill = PatternFill(start_color="AFEF61", end_color="AFEF61", fill_type="solid")

        # 写入厚度数据
        for col, thickness in enumerate(thickness_data, start=2):
            for row, value in enumerate(thickness, start=2):
                ws.cell(row=row, column=col, value=value)

        # 保存 Excel 文件
        wb.save(file_path)

    def process(self):
        self.unified_sequence = self.get_unified_sequence()
        self.standardized_boreholes = self.complete_missing_layers()
        self.preconditioning_completes_borehole = self.make_max_compatible_boreholes()
        self.print_pandas(deepcopy(self.preconditioning_completes_borehole))
        self.boreholes_thickness = self.zero_thickness_filling()
        self.print_pandas(deepcopy(self.boreholes_thickness))
        # 调用写入 Excel 方法
        self.write_to_excel("borehole_data.xlsx")

def execute_main():
    boreholes = {
        "BH01": ["填土", "黄土", "粉土", "粉砂", "砂岩"],
        "BH02": ["填土", "黄土", "粉砂", "砂岩", "泥岩"],
        "BH03": ["填土", "粉土", "黄土", "砂岩", "砾石"],
        "BH04": ["填土", "黄土", "粉土", "泥岩", "页岩"],
        "BH05": ["填土", "黄土", "粉砂", "页岩", "黏土"],
        "BH07": ["填土", "砂岩", "粉砂", "泥岩", "砾石"],
        "BH08": ["填土", "粉土", "黄土", "粉砂", "泥岩"],
        "BH09": ["填土", "砂岩", "砾石", "页岩", "粉土"],
        "BH10": ["填土", "粉砂", "黄土", "粉砂", "砂岩"],
    }
    boreholes_thickness = {
        "BH01": [11, 13, 14, 16, 20],
        "BH02": [30, 24, 25, 13, 31],
        "BH03": [23, 14, 15, 26, 26],
        "BH04": [13, 14, 15, 15, 16],
        "BH05": [34, 25, 25, 26, 16],
        "BH07": [13, 14, 15, 15, 26],
        "BH08": [12, 13, 14, 15, 15],
        "BH09": [22, 21, 21, 21, 21],
        "BH10": [24, 24, 24, 24, 24],
    }
    location = {
        "BH01": [11, 13],
        "BH02": [30, 24],
        "BH03": [23, 14],
        "BH04": [13, 14],
        "BH05": [34, 25],
        "BH07": [13, 14],
        "BH08": [12, 13],
        "BH09": [22, 21],
        "BH10": [24, 24],
    }
    processor = boreholeProcessor(boreholes, boreholes_thickness)
    processor.process()

if __name__ == "__main__":
    execute_main()







